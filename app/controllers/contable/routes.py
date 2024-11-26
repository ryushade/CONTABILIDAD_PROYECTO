from app.models.conexion import obtener_conexion
from flask import json, Response, abort, request, redirect, url_for, flash, session, jsonify, render_template, send_file, current_app, send_from_directory
from flask_jwt_extended import jwt_required, create_access_token, set_access_cookies, unset_jwt_cookies, get_jwt_identity, verify_jwt_in_request

from app.models.contable_models import actualizar_regla_en_db, obtener_numero_reglas_por_tipo_transaccion, agregar_regla_en_db, obtener_id_cuenta , cuentas_jerarquicas, guardar_foto_usuario, obtener_regla_por_id, obtener_roles, obtener_usuario_por_nombre, agregar_usuario, actualizar_usuario, eliminar_usuario, verificar_contraseña, obtener_asientos_agrupados,obtener_reglas, obtener_cuentas, obtener_usuarios, obtener_total_cuentas, eliminar_cuenta_contable, eliminar_regla_bd, obtener_usuario_por_id, obtener_cuenta_por_id, actualizar_cuenta, obtener_cuentas_excel, obtener_libro_mayor_agrupado_por_fecha, obtener_libro_mayor_agrupado_por_fecha_y_glosa_unica,obtener_registro_ventas, obtener_asientos_agrupados_excel, obtener_libro_caja, obtener_libro_caja_cuenta_corriente, actualizar_rol_usuario, obtener_registro_compras
from functools import wraps
from flask import current_app as app
from . import accounting_bp
import pandas as pd
import io
import os
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from openpyxl.styles import Border, Side, Alignment, Font
from fpdf import FPDF
from io import BytesIO
from werkzeug.utils import secure_filename
from flask_jwt_extended import get_jwt_identity
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import numbers 
import pdfkit
from weasyprint import HTML, CSS
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def nocache(view):
    @wraps(view)
    def no_cache(*args, **kwargs):
        response = make_response(view(*args, **kwargs))
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    return no_cache


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Verificar el JWT
            verify_jwt_in_request()
            user_id = get_jwt_identity()
            user = obtener_usuario_por_id(user_id)
            if user and user['rol']['nom_rol'] in roles:
                return f(*args, **kwargs)
            else:
                # Redirige al inicio con un parámetro para indicar acceso denegado
                return redirect(url_for('inicio', acceso_denegado=1))
        return decorated_function
    return decorator

from flask import Flask, request, redirect, url_for, render_template, flash, make_response

@accounting_bp.route('/login', methods=['GET', 'POST'])
@nocache
def login():
    if request.method == 'GET':
        # Invalida el token JWT y limpia la sesión
        response = make_response(render_template('contable/login.html'))
        unset_jwt_cookies(response)
        session.clear()
        return response

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Consulta el usuario por nombre
        user = obtener_usuario_por_nombre(username)

        # Verifica si el usuario existe y la contraseña es correcta
        if user and verificar_contraseña(user, password):
            # Genera el token de acceso JWT
            access_token = create_access_token(identity=user['id_usuario'])
            response = make_response(redirect(url_for('inicio')))
            set_access_cookies(response, access_token)
            session['username'] = username
            response.set_cookie('username', username, httponly=True)  # Guarda el username en una cookie
            session['username'] = username
            print(session.get('username'))
            return response
        else:
            # Si el inicio de sesión falla, redirige con un mensaje de error
            return redirect(url_for('contable.login') + '?error=true')



@accounting_bp.route('/logout')
def logout():
    response = make_response(redirect(url_for('contable.login')))
    unset_jwt_cookies(response)  # Elimina las cookies JWT
    response.set_cookie('username', '', expires=0)  # Elimina la cookie del nombre de usuario
    session.clear()  # Limpia la sesión
    return response

@accounting_bp.route('/cambiar_rol', methods=['POST'])
@jwt_required()
def cambiar_rol():
    user_id = get_jwt_identity()
    user = obtener_usuario_por_id(user_id)
    
    if not user or not user.get('admin'):
        return redirect(url_for('inicio'))

    selected_role_id = request.form.get('selected_role')
    if not selected_role_id:
        return redirect(url_for('inicio'))

    # Validar que el rol seleccionado es válido
    allowed_roles = [1, 2, 3]  # ADMIN, CONTADOR, EMPLEADO
    try:
        selected_role_id = int(selected_role_id)
    except ValueError:
        return redirect(url_for('inicio'))

    if selected_role_id not in allowed_roles:
        return redirect(url_for('inicio'))

    # Actualizar el id_rol del usuario
    try:
        actualizar_rol_usuario(user_id, selected_role_id)
    except Exception as e:
        app.logger.error(f"Error al actualizar el rol: {str(e)}")

    print(f"Usuario ID: {user_id}")
    print(f"Rol seleccionado: {selected_role_id}")

    return redirect(url_for('inicio'))

@accounting_bp.route('/reportes', methods=['GET'])
@jwt_required()
@role_required('ADMIN', 'CONTADOR')
def reportes():
    active_tab = request.args.get('active_tab', 'libro-diario')
    tipo_registro = request.args.get('tipo_registro', 'Todas')
    daterange = request.args.get('daterange', '')
    daterange_mayor = request.args.get('daterangemayor', '')
    daterange_caja = request.args.get('daterangecaja', '')
    daterange_venta = request.args.get('daterangeventa', '')
    daterange_compra = request.args.get('daterangecompra', '')
    start_date, end_date = None, None
    start_date_mayor, end_date_mayor = None, None
    start_date_caja, end_date_caja = None, None
    start_date_venta, end_date_venta = None, None
    start_date_compra, end_date_compra = None, None

    def last_day_of_month(any_day):
        next_month = any_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def parse_daterange(daterange):
        try:
            # Limpia espacios adicionales y remueve espacios al inicio y fin
            daterange_clean = re.sub(r'\s+', ' ', daterange.strip())
            # Divide por ' to ' o ' a '
            dates = re.split(r'\s*(to|a)\s*', daterange_clean)
            # Remueve los separadores del resultado
            dates = [d for d in dates if d not in ('to', 'a')]
            if len(dates) == 2:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = datetime.strptime(dates[1].strip(), '%m/%Y').date()
                end_date = last_day_of_month(end_date)
            elif len(dates) == 1:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = last_day_of_month(start_date)
            else:
                start_date = end_date = None
            return start_date, end_date
        except (ValueError, IndexError) as e:
            print(f"Error processing date range: {daterange} - {e}")
            return None, None

    # Procesar los rangos de fechas utilizando la función parse_daterange
    if daterange:
        start_date, end_date = parse_daterange(daterange)

    if daterange_mayor:
        start_date_mayor, end_date_mayor = parse_daterange(daterange_mayor)

    if daterange_caja:
        start_date_caja, end_date_caja = parse_daterange(daterange_caja)

    if daterange_venta:
        start_date_venta, end_date_venta = parse_daterange(daterange_venta)

    if daterange_compra:
        start_date_compra, end_date_compra = parse_daterange(daterange_compra)

    # Llamar a las funciones para obtener los datos, pasando las fechas procesadas
    asientos, totales = obtener_asientos_agrupados(tipo_registro, start_date, end_date)
    libro_mayor_data, total_debe, total_haber = obtener_libro_mayor_agrupado_por_fecha(start_date_mayor, end_date_mayor)
    lista_libro_caja, total_caja = obtener_libro_caja(start_date_caja, end_date_caja)
    lista_libro_caja_cuenta_corriente, total_caja_corriente = obtener_libro_caja_cuenta_corriente()
    registro_venta_data, totale = obtener_registro_ventas(start_date_venta, end_date_venta)
    registro_compra_data, totales_compra = obtener_registro_compras(start_date_compra, end_date_compra)

    return render_template(
        'contable/reportes/reportes.html', 
        asientos=asientos, 
        totales=totales, 
        libro_mayor=libro_mayor_data, 
        total_debe=total_debe, 
        total_haber=total_haber, 
        registros_ventas=registro_venta_data, 
        totale=totale, 
        registros_compras=registro_compra_data,
        totalesCompras=totales_compra,
        lista_libro_caja=lista_libro_caja, 
        total_caja=total_caja, 
        lista_libro_caja_cuenta_corriente=lista_libro_caja_cuenta_corriente, 
        total_caja_corriente=total_caja_corriente,
        active_tab=active_tab
    )

def last_day_of_month(any_day):
    next_month = any_day.replace(day=28) + timedelta(days=4)
    return next_month - timedelta(days=next_month.day)

def parse_daterange(daterange):
    try:
        # Limpia espacios adicionales y remueve espacios al inicio y fin
        daterange_clean = re.sub(r'\s+', ' ', daterange.strip())
        # Divide por ' to ' o ' a '
        dates = re.split(r'\s*(to|a)\s*', daterange_clean)
        # Remueve los separadores del resultado
        dates = [d for d in dates if d not in ('to', 'a')]
        if len(dates) == 2:
            start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
            end_date = datetime.strptime(dates[1].strip(), '%m/%Y').date()
            end_date = last_day_of_month(end_date)
        elif len(dates) == 1:
            start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
            end_date = last_day_of_month(start_date)
        else:
            start_date = end_date = None
        return start_date, end_date
    except (ValueError, IndexError) as e:
        print(f"Error processing date range: {daterange} - {e}")
        return None, None
    

from flask import get_flashed_messages
import app.models.contable_models as conta
@accounting_bp.route('/cuentas', methods=['GET'])
@jwt_required()
@role_required('ADMIN', 'CONTADOR')
def cuentas():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    tipo_cuenta = request.args.get('tipo_cuenta')
    naturaleza = request.args.get('naturaleza')
    cuentas = conta.obtenerCuentas()
    # print(json.dumps(estructura_cuentas, indent=2))
    total_cuentas = obtener_total_cuentas(tipo_cuenta=tipo_cuenta, naturaleza=naturaleza)
    total_pages = (total_cuentas + per_page - 1) // per_page
    # Obtener el mensaje de error
    error_message = session.pop('error_message', None)
    error_eliminar = get_flashed_messages(category_filter=['error'])
    success_messages = get_flashed_messages(category_filter=['success'])
    print(error_message)
    return render_template(
        'contable/cuentas/cuentas2.html',
        cuentas = cuentas,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        total_results=total_cuentas,
        tipo_cuenta=tipo_cuenta,
        naturaleza=naturaleza,
        error_message = error_message,
        max=max,
        min=min,
        error_eliminar = error_eliminar,
        success_messages = success_messages
    )


@accounting_bp.route('/exportar_excel', methods=['GET'])
@jwt_required()
@role_required('ADMIN', 'CONTADOR')
def exportar_excel():

    cuentas = cuentas_jerarquicas()


    df = pd.DataFrame(cuentas)


    df.rename(columns={
        'codigo_cuenta': 'Código de Cuenta',
        'nombre_cuenta': 'Nombre de Cuenta',
        'cuenta_padre': 'Cuenta Padre',
        'nivel': 'Nivel',
        'jerarquia': 'Jerarquía'
    }, inplace=True)

   
    df['Elemento'] = (df['Nivel'] == 1).cumsum()

 
    def formatear_nombre(row):
        if row['Nivel'] == 1:
            nombre_formateado = f'ELEMENTO {row["Elemento"]}: {row["Nombre de Cuenta"]}'
        else:
            nombre_formateado = row['Nombre de Cuenta']
        
        indentacion = '    ' * (row['Nivel'] - 1)
        return f'{indentacion}{nombre_formateado}'


    df['Nombre de Cuenta Formateado'] = df.apply(formatear_nombre, axis=1)

    
    df = df[['Código de Cuenta', 'Nombre de Cuenta Formateado', 'Nivel']]

    
    df.rename(columns={'Nombre de Cuenta Formateado': 'Nombre de Cuenta'}, inplace=True)

    
    output = io.BytesIO()

    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Plan de Cuentas', index=False)

        
        workbook = writer.book
        worksheet = writer.sheets['Plan de Cuentas']

        # Definir estilos para el encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar estilos al encabezado
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet[f'{get_column_letter(col_num)}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Ajustar el ancho de las columnas según el contenido
        for col_num, column in enumerate(df.columns, 1):
            max_length = max(
                df[column].astype(str).map(len).max(),
                len(column)
            ) + 2  # Añadir un poco de espacio adicional
            worksheet.column_dimensions[get_column_letter(col_num)].width = max_length

        # Aplicar estilos condicionales a las filas
        for row in range(2, worksheet.max_row + 1):
            nivel = df.at[row - 2, 'Nivel'] 
            nombre_cuenta_cell = worksheet[f'B{row}'] 

            if nivel == 1 or nivel == 2:
                nombre_cuenta_cell.font = Font(bold=True)
            else:
                nombre_cuenta_cell.font = Font(bold=False)

            # Aplicar borde a todas las celdas de la fila
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col_num)
                cell.border = thin_border

    
    output.seek(0)

    return send_file(
        output,
        download_name="plan_de_cuentas.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@accounting_bp.route('/exportar_pdf', methods=['GET'], endpoint='exportar_pdf')
@jwt_required()
@role_required('ADMIN', 'CONTADOR')
def descargar_pdf():
    try:
        # Obtener las cuentas jerárquicas
        cuentas = cuentas_jerarquicas()

        # Renderizar el HTML con la plantilla
        rendered_html = render_template('contable/cuentas/pcge.html', cuentas=cuentas)

        # Generar el PDF con WeasyPrint
        pdf = HTML(string=rendered_html).write_pdf(
            stylesheets=[CSS("app/static/css/cuenta.css")],
            presentational_hints=True
        )

        return Response(
            pdf,
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment; filename="Plan_Contable.pdf"'}
        )
    except Exception as e:
        return f"Error al generar el PDF: {e}", 500

@accounting_bp.route('/cuentas/obtener/<int:cuenta_id>', methods=['GET'])
def obtener_cuenta(cuenta_id):
    cuenta = obtener_cuenta_por_id(cuenta_id)
    if cuenta:
        return jsonify(cuenta)
    else:
        return jsonify({'error': 'Cuenta no encontrada'}), 404
    


@accounting_bp.route('/cuentas/editar/<int:cuenta_id>', methods=['POST'])
def editar_cuenta(cuenta_id):
    codigo_cuenta = request.form['codigo_cuenta']
    nombre_cuenta = request.form['nombre_cuenta']
    naturaleza = request.form['naturaleza']
    estado_cuenta = request.form['estado_cuenta']

    actualizar_cuenta(cuenta_id, codigo_cuenta, nombre_cuenta, naturaleza, estado_cuenta)

    return redirect(url_for('contable.cuentas'))

@accounting_bp.route('/cuentas/registrar', methods=['POST'])
def registrar_cuenta():
    try:
        codigo_cuenta = request.form['codigo_cuenta_agregar']
        nombre_cuenta = request.form['nombre_cuenta_agregar']
        naturaleza = request.form['naturaleza_agregar']
        estado_cuenta = request.form['estado_cuenta_agregar']
        print(codigo_cuenta)
        print(nombre_cuenta)
        print(naturaleza)
        print(estado_cuenta)
        conta.insertar_cuenta(codigo_cuenta, nombre_cuenta, naturaleza, estado_cuenta)
        return redirect(url_for('contable.cuentas'))
    except Exception as e:
        session['error_message'] = str(e)
        return redirect(url_for('contable.cuentas'))


@accounting_bp.route('/reglas/agregar', methods=['POST'])
def agregar_regla():
    data = request.get_json()

    nombre_regla = data.get("nombre_regla")
    tipo_transaccion = data.get("tipo_transaccion")
    cuenta_debito_codigo = data.get("cuenta_debito") or None
    cuenta_credito_codigo = data.get("cuenta_credito") or None
    estado = data.get("estado")
    tipo_monto = data.get("tipo_monto")

    print("Datos recibidos en agregar_regla:")
    print("Nombre de la regla:", nombre_regla)
    print("Tipo de transacción:", tipo_transaccion)
    print("Cuenta débito código:", cuenta_debito_codigo)
    print("Cuenta crédito código:", cuenta_credito_codigo)
    print("Estado:", estado)

    if tipo_transaccion is None:
        print("Error: tipo_transaccion es None")
        return jsonify({"success": False, "message": "El tipo de transacción no puede ser None"}), 400

    # Convertir los códigos de cuenta a id_cuenta antes de agregar la regla
    cuenta_debito = obtener_id_cuenta(cuenta_debito_codigo) if cuenta_debito_codigo else None
    cuenta_credito = obtener_id_cuenta(cuenta_credito_codigo) if cuenta_credito_codigo else None

    print("Cuenta Débito ID:", cuenta_debito)
    print("Cuenta Crédito ID:", cuenta_credito)

    try:
        # Verificar el número de reglas existentes para el tipo de transacción
        numero_reglas = obtener_numero_reglas_por_tipo_transaccion(tipo_transaccion)
        if numero_reglas >= 3:
            return jsonify({
                "success": False,
                "message": "No se pueden agregar más de 3 reglas para este tipo de transacción."
            }), 400

        resultado = agregar_regla_en_db(
            nombre_regla,
            tipo_transaccion,
            cuenta_debito,
            cuenta_credito,
            estado,
            tipo_monto
        )

        if resultado:
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "message": "No se pudo agregar la regla."}), 400

    except Exception as e:
        print("Error al agregar la regla:", e)
        return jsonify({"success": False, "message": "Error interno del servidor."}), 500



@accounting_bp.route('/upload_photo', methods=['POST'])
def upload_photo():
    if 'photo' not in request.files:
        return redirect(url_for('inicio'))

    file = request.files['photo']
    if file.filename == '':
        return redirect(url_for('inicio'))

    username = request.cookies.get('username', None)
    print(username)
    if not username:
        return redirect(url_for('inicio'))  # Si no hay cookie, redireccionar al inicio

    filename = secure_filename(username)
    if not filename:
        return redirect(url_for('inicio'))

    file_extension = os.path.splitext(file.filename)[1]
    filename += file_extension

    # Asegura que estás usando el directorio correcto
    upload_folder = current_app.config['UPLOAD_FOLDER']
    print("upload_folder", upload_folder)
    # Crea la carpeta si no existe
    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)

    filepath = os.path.join(upload_folder, filename)
    file.save(filepath)

    # Genera la ruta para almacenar en la base de datos
    foto_path = f"/app/static/fotos_perfil/{filename}"
    print(f"Ruta URL para la base de datos: {foto_path}")

    user_id = request.form.get('user_id')
    if not user_id:
        return redirect(url_for('inicio'))

    # Guardar la ruta en la base de datos
    if guardar_foto_usuario(user_id, foto_path):
        return redirect(url_for('inicio'))
    else:
        return redirect(url_for('inicio'))
    
@accounting_bp.route('/upload_photo/obtener/<int:user_id>', methods=['GET'])
def obtener_foto_usuario(user_id):
    foto = obtener_foto_usuario_db(user_id)
    def obtener_foto_usuario_db(user_id):
        conexion = obtener_conexion()
        try:
            with conexion.cursor() as cursor:
                query = "SELECT foto FROM usuario WHERE id_usuario = %s"
                cursor.execute(query, (user_id,))
                result = cursor.fetchone()
                if result:
                    return result['foto']
                else:
                    return None
        except Exception as e:
            print(f"Error al obtener la foto del usuario: {e}")
            return None
        finally:
            conexion.close()
    if foto:
        return jsonify({"foto_path": foto})
    else:
        return jsonify({"foto_path": None})
    

@accounting_bp.route('/reglas/actualizar_regla/<int:id_regla>', methods=['POST'])
def actualizar_regla(id_regla):
    data = request.get_json()

    nombre_regla = data.get("nombre_regla")
    tipo_transaccion = data.get("tipo_transaccion")
    cuenta_debito_codigo = data.get("cuenta_debito") or None
    cuenta_credito_codigo = data.get("cuenta_credito") or None
    estado = data.get("estado")

    # Añade estos prints
    print("Datos recibidos:")
    print("ID Regla:", id_regla)
    print("Nombre Regla:", nombre_regla)
    print("Tipo Transacción:", tipo_transaccion)
    print("Cuenta Débito Código:", cuenta_debito_codigo)
    print("Cuenta Crédito Código:", cuenta_credito_codigo)
    print("Estado:", estado)

       # Convertir los códigos de cuenta a id_cuenta antes de actualizar
    cuenta_debito = obtener_id_cuenta(cuenta_debito_codigo) if cuenta_debito_codigo else None
    cuenta_credito = obtener_id_cuenta(cuenta_credito_codigo) if cuenta_credito_codigo else None

    print("Cuenta Débito ID:", cuenta_debito)
    print("Cuenta Crédito ID:", cuenta_credito)

    resultado = actualizar_regla_en_db(id_regla, nombre_regla, tipo_transaccion, cuenta_debito, cuenta_credito, estado)

    if resultado:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "message": "No se encontró la regla o no se pudo actualizar."}), 404

@accounting_bp.route('/reglas', methods=['GET'])
@jwt_required()
@role_required('ADMIN', 'CONTADOR')
def reglas():
    page = request.args.get('page', default=1, type=int)
    per_page = request.args.get('per_page', default=5, type=int)
    tipo_transaccion = request.args.get('tipo_transaccion', default='Todas')
    tipo_monto = request.args.get('tipo_monto', default='Todas')

    # Llamar a la función para obtener las reglas y el total de resultados
    reglas, total_results = obtener_reglas(page, per_page, tipo_transaccion, tipo_monto)

    # Calcular el total de páginas
    total_pages = (total_results + per_page - 1) // per_page

    # Renderizar la plantilla
    return render_template(
        'contable/reglas/reglas.html',
        reglas=reglas,
        page=page,
        per_page=per_page,
        total_results=total_results,
        total_pages=total_pages,
        tipo_transaccion=tipo_transaccion,
        tipo_monto=tipo_monto,
        max=max,
        min=min,
    )




@accounting_bp.route('/usuarios', methods=['GET'])
@jwt_required()
@role_required('ADMIN')
def usuarios():
    page = request.args.get('page', default=1, type=int)
    per_page = request.args.get('per_page', default=5, type=int)

    usuarios, total_results = obtener_usuarios(page, per_page)
    roles = obtener_roles()

    total_pages = (total_results + per_page - 1) // per_page

    return render_template(
        'contable/usuarios/usuarios.html',
        usuarios=usuarios,
        roles=roles,
        page=page,
        per_page=per_page,
        total_results=total_results,
        total_pages=total_pages,
        max=max,  
        min=min   
    )


@accounting_bp.route('/usuarios/obtener/<int:id_usuario>', methods=['GET'])
def obtener_usuario(id_usuario):
    conexion = obtener_conexion()
    try:
        with conexion.cursor() as cursor:
            # Consulta para obtener datos del usuario y el nombre del rol
            query = """
                SELECT usuario.id_rol, usuario.usua, usuario.contra, usuario.estado_usuario, rol.nombre_rol 
                FROM usuario 
                JOIN rol ON usuario.id_rol = rol.id_rol
                WHERE usuario.id_usuario = %s
            """
            cursor.execute(query, (id_usuario,))
            usuario = cursor.fetchone()

            if usuario:
                return jsonify({
                    "id_rol": usuario[0],
                    "usua": usuario[1],
                    "contra": usuario[2],
                    "estado_usuario": usuario[3],
                    "nombre_rol": usuario[4]  # Nombre del rol
                })
            else:
                return jsonify({"error": "Usuario no encontrado"}), 404
    except Exception as error:
        return jsonify({"error": str(error)}), 500
    finally:
        conexion.close()


@accounting_bp.route('/reglas/detalles/<int:regla_id>', methods=['GET'])
def obtener_detalles_regla(regla_id):
    regla = obtener_regla_por_id(regla_id)
    if regla:
        return jsonify({
            "nombre_regla": regla.get("nombre_regla"),
            "tipo_transaccion": regla.get("tipo_transaccion"),
            "estado": "Activo" if regla.get("estado") == 1 else "Inactivo",
            "cuenta_debito_codigo": regla.get("cuenta_debe_codigo"),
            "cuenta_debito_nombre": regla.get("cuenta_debe_nombre"),
            "cuenta_credito_codigo": regla.get("cuenta_haber_codigo"),
            "cuenta_credito_nombre": regla.get("cuenta_haber_nombre"),
            "tipo_monto" : regla.get("tipo_monto")
        })
    else:
        return jsonify({'error': 'Regla no encontrada'}), 404




@accounting_bp.route('/usuarios/actualizar/<int:id_usuario>', methods=['POST'])
def actualizar_usu(id_usuario):
    id_rol = request.form.get('id_rol')
    usua = request.form.get('usua')
    contra = request.form.get('contrasena')
    estado_usuario = request.form.get('estado')
    admin = request.form.get('admin')

    print("Valores recibidos en el servidor:")
    print(f"id_rol: {id_rol}, usua: {usua}, contra: {contra}, estado: {estado_usuario}, admin: {admin}")

    resultado = actualizar_usuario(id_usuario, id_rol, usua, contra, estado_usuario, admin)
    print("Resultado de la actualización:", resultado)
    if "error" in resultado:
        return jsonify(resultado), 400

    return jsonify(resultado), 200


@accounting_bp.route('/cuentas/pcge', methods=['GET'])
@jwt_required()
def prueba_pcge():
    cuentas = cuentas_jerarquicas()  # Obtener las cuentas distribuidas
    return render_template('contable/cuentas/pcge.html', cuentas=cuentas)  # Pasa las columnas al template


@accounting_bp.route('/usuarios/agregar', methods=['POST'])
def agregar_usu():
    datos = request.get_json()
    id_rol = datos.get('id_rol')
    usua = datos.get('usua')
    contra = datos.get('contra')
    estado_usuario = datos.get('estado_usuario')
    admin = datos.get('admin')

    # Verifica que los datos no sean nulos
    if not id_rol or not usua or not contra or not estado_usuario:
        return jsonify({"message": "Bad Request. Please fill all fields."}), 400

    resultado = agregar_usuario(id_rol, usua, contra, estado_usuario, admin)

    # Verifica si hubo un error y devuelve un mensaje apropiado
    if "error" in resultado:
        return jsonify({"code": 0, "message": resultado["error"]}), 400

    return jsonify({"code": 1, "message": "Usuario añadido"})


@accounting_bp.route('/usuarios/eliminar/<int:usuario_id>', methods=['POST'])
def eliminar_usu(usuario_id):
    usuario_a_eliminar = obtener_usuario_por_id(usuario_id)
    
    if not usuario_a_eliminar:
        flash("El usuario no existe o ya ha sido eliminado.", "warning")
        return redirect(url_for('contable.usuarios'))

    eliminar_usuario(usuario_id)

    if 'username' in session and session.get('username') == usuario_a_eliminar['usua']:
        session.clear()
        flash("Tu cuenta ha sido eliminada. Por favor, inicia sesión nuevamente.", "warning")
        
        response = redirect(url_for('contable.login'))
        unset_jwt_cookies(response)
        return response
    
    flash("Usuario eliminado exitosamente.", "success")
    return redirect(url_for('contable.usuarios'))



@accounting_bp.route('/cuentas/eliminar/<int:cuenta_id>', methods=['POST'])
def eliminar_cuenta(cuenta_id):
    try:
        eliminar_cuenta_contable(cuenta_id)
        flash("Cuenta eliminada correctamente", category='success')
        return redirect(url_for('contable.cuentas'))
    except Exception as e:
        flash(str(e), category='error')  # Asegúrate de especificar una categoría adecuada
        return redirect(url_for('contable.cuentas'))

@accounting_bp.route('/reglas/eliminar/<int:regla_id>', methods=['POST'])
def eliminar_regla(regla_id):
    eliminar_regla_bd(regla_id)  # Llama a la función separada
    return redirect(url_for('contable.reglas'))

@accounting_bp.route('/reportes/ldpdf')
@jwt_required()
def ldpf():
    return render_template('contable/reportes/pdf_ld.html')

@accounting_bp.route('/exportar_libro_diario_excel', methods=['GET'])
@jwt_required()
def exportar_libro_diario_excel():
    import re  # Asegúrate de importar 're' para expresiones regulares
    import openpyxl
    import os
    from flask import current_app, send_file, request
    from io import BytesIO
    from datetime import datetime, timedelta
    from openpyxl.styles import Alignment
    template_path = os.path.join(current_app.root_path, 'templates', 'contable', 'plantillas', 'L,D.xlsx')
    
    if not os.path.exists(template_path):
        return jsonify({'error': 'La plantilla de Excel no se encontró.'}), 404
    
    output = BytesIO()
    
    workbook = load_workbook(template_path)
    worksheet = workbook.active
    
    # Capturar filtros desde la URL
    tipo_registro = request.args.get('tipo_registro', 'Todas')
    daterange = request.args.get('daterange', '')
    start_date, end_date = None, None

    # Procesar rango de fechas y determinar el mes para B3
    mes_anio_excel = None  # Variable para guardar el valor de B3

    # Función para procesar el rango de fechas (puedes extraerla a un módulo común si lo prefieres)
    def parse_daterange(daterange):
        try:
            daterange_clean = re.sub(r'\s+', ' ', daterange.strip())
            dates = re.split(r'\s*(to|a)\s*', daterange_clean)
            dates = [d for d in dates if d not in ('to', 'a')]
            if len(dates) == 2:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = datetime.strptime(dates[1].strip(), '%m/%Y').date()
                # Calcular el último día del mes para el segundo mes
                end_date = end_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y') + " - " + end_date.strftime('%m/%Y')
            elif len(dates) == 1:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                # Calcular el último día del mes para el mes seleccionado
                end_date = start_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y')
            else:
                start_date = end_date = None
            return start_date, end_date, mes_anio_excel
        except (ValueError, IndexError) as e:
            print(f"Error processing date range: {daterange} - {e}")
            return None, None, None

    # Procesar el rango de fechas
    if daterange:
        start_date, end_date, mes_anio_excel = parse_daterange(daterange)

    # Si no se proporcionó un rango de fechas, usar el mes actual
    if not mes_anio_excel:
        fecha_actual = datetime.now()
        mes_anio_excel = fecha_actual.strftime('%m/%Y')
        start_date = fecha_actual.replace(day=1).date()
        end_date = (fecha_actual.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
    
    # Llenar las celdas específicas con los datos necesarios
    worksheet['B3'] = mes_anio_excel  
    worksheet['B4'] = '20610588981'
    worksheet['B5'] = 'Tormenta'
    
    # Definir la fila inicial para insertar los datos en la tabla
    start_row = 11

    # Obtener los asientos aplicando los filtros
    asientos, _ = obtener_asientos_agrupados(tipo_registro, start_date, end_date)

    # Verificar si hay datos para exportar
    if not asientos:
        response = make_response(jsonify({'error': 'No hay datos para exportar en el rango de fechas seleccionado.'}), 200)
        response.headers['Content-Type'] = 'application/json'
        return response

    numero_correlativo = 1

    total_debe = 0
    total_haber = 0

    # Crear un borde negro
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Crear un estilo de fuente sin negrita
    normal_font = Font(bold=False)

    # Estilos de alineación
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Insertar los datos en la tabla, fila por fila
    current_row = start_row
    for id_asiento, asiento in asientos.items():
        
        # Guardar la fila inicial del grupo para combinar celdas
        start_group_row = current_row

        # Extraer los dígitos del número de comprobante para la columna E
        num_correlativo_lm = asiento['num_comprobante'][1:4]  # Extraer los caracteres en posición 2, 3 y 4

        for detalle in asiento['detalles']:

            worksheet[f'A{current_row}'].number_format = numbers.FORMAT_TEXT
            worksheet[f'A{current_row}'] = str(numero_correlativo)
            worksheet[f'B{current_row}'] = asiento['fecha_asiento'].strftime('%d/%m/%Y')
            worksheet[f'C{current_row}'] = asiento['glosa']
            
            # Determinar el valor para la columna D
            if 'venta' in asiento['glosa'].lower() or 'ingreso' in asiento['glosa'].lower():
                worksheet[f'D{current_row}'] = 14
            elif 'compra' in asiento['glosa'].lower():
                worksheet[f'D{current_row}'] = 8
            elif 'pago' in asiento['glosa'].lower():
                worksheet[f'D{current_row}'] = 1
            else:
                worksheet[f'D{current_row}'] = ''
            
            worksheet[f'F{current_row}'] = asiento['num_comprobante']
            worksheet[f'G{current_row}'] = detalle['codigo_cuenta']
            worksheet[f'H{current_row}'] = detalle['nombre_cuenta']
            worksheet[f'I{current_row}'] = detalle['debe'] if detalle['debe'] != 0 else None
            worksheet[f'J{current_row}'] = detalle['haber'] if detalle['haber'] != 0 else None

            # Aplicar formato de número con dos decimales a las columnas "Debe" y "Haber"
            worksheet[f'I{current_row}'].number_format = '#,##0.00'
            worksheet[f'J{current_row}'].number_format = '#,##0.00'

            # Aplicar estilo a todas las celdas
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                cell = worksheet[f'{col}{current_row}']
                cell.border = thin_border
                cell.font = normal_font

            # Alineación específica para columnas
            worksheet[f'A{current_row}'].alignment = center_alignment
            worksheet[f'B{current_row}'].alignment = left_alignment
            worksheet[f'C{current_row}'].alignment = left_alignment
            worksheet[f'D{current_row}'].alignment = center_alignment
            worksheet[f'F{current_row}'].alignment = left_alignment
            worksheet[f'G{current_row}'].alignment = center_alignment
            worksheet[f'H{current_row}'].alignment = left_alignment
            worksheet[f'I{current_row}'].alignment = right_alignment
            worksheet[f'J{current_row}'].alignment = right_alignment
            
            total_debe += detalle['debe']
            total_haber += detalle['haber']

            current_row += 1

        # Combinar celdas para las columnas indicadas
        if current_row > start_group_row + 1:  # Combinar solo si hay más de una fila en el grupo
            worksheet.merge_cells(f'A{start_group_row}:A{current_row - 1}')
            worksheet.merge_cells(f'B{start_group_row}:B{current_row - 1}')
            worksheet.merge_cells(f'C{start_group_row}:C{current_row - 1}')
            worksheet.merge_cells(f'D{start_group_row}:D{current_row - 1}')
            worksheet.merge_cells(f'E{start_group_row}:E{current_row - 1}')
            worksheet.merge_cells(f'F{start_group_row}:F{current_row - 1}')
            
            # Alinear las celdas combinadas
            worksheet[f'A{start_group_row}'].alignment = center_alignment
            worksheet[f'B{start_group_row}'].alignment = left_alignment
            worksheet[f'C{start_group_row}'].alignment = left_alignment
            worksheet[f'D{start_group_row}'].alignment = center_alignment
            worksheet[f'E{start_group_row}'].alignment = center_alignment
            worksheet[f'F{start_group_row}'].alignment = left_alignment

        # Escribir el valor en la columna E (una vez por grupo de filas)
        worksheet[f'E{start_group_row}'] = num_correlativo_lm

        numero_correlativo += 1
    
    # Combinar las celdas de la fila del total
    total_row = current_row
    worksheet.merge_cells(f'A{total_row}:H{total_row}')
    worksheet[f'A{total_row}'] = "Total"
    worksheet[f'A{total_row}'].alignment = Alignment(horizontal='right')
    worksheet[f'A{total_row}'].border = thin_border

    worksheet[f'I{total_row}'] = total_debe
    worksheet[f'J{total_row}'] = total_haber

    # Aplicar formato de número con dos decimales a los totales
    worksheet[f'I{total_row}'].number_format = '#,##0.00'
    worksheet[f'J{total_row}'].number_format = '#,##0.00'

    # Alinear a la derecha los totales de Debe y Haber
    worksheet[f'I{total_row}'].alignment = right_alignment
    worksheet[f'J{total_row}'].alignment = right_alignment
    
    for col in ['I', 'J']:
        cell = worksheet[f'{col}{total_row}']
        cell.border = thin_border
        cell.font = normal_font

    # Asegurar que todas las columnas tengan un ancho adecuado
    column_widths = {'A': 40, 'B': 15, 'C': 65, 'D': 10, 'F': 20, 'G': 15, 'H': 40, 'I': 15, 'J': 15}
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    workbook.save(output)
    output.seek(0)

    return send_file(output, download_name="libro_diario.xlsx", as_attachment=True)

from flask import current_app, send_file, request, jsonify, make_response
from io import BytesIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment
import os
import re

@accounting_bp.route('/exportar_libro_mayor_excel', methods=['GET'])
@jwt_required()
def exportar_libro_mayor_excel():
    # Ruta de la plantilla
    template_path = os.path.join(current_app.root_path, 'templates', 'contable', 'plantillas', '234_formato51.xlsx')
    
    if not os.path.exists(template_path):
        return jsonify({'error': 'La plantilla de Excel no se encontró.'}), 404
    
    # Capturar filtros desde la URL
    daterangemayor = request.args.get('daterangemayor', '')
    start_date, end_date = None, None

    mes_anio_excel = None  # Variable para guardar el valor de B3

    # Función para procesar el rango de fechas
    def parse_daterange(daterange):
        try:
            # Limpiar espacios adicionales
            daterange_clean = re.sub(r'\s+', ' ', daterange.strip())
            # Dividir por ' to ' o ' a '
            dates = re.split(r'\s*(to|a)\s*', daterange_clean)
            # Remover los separadores del resultado
            dates = [d for d in dates if d not in ('to', 'a')]
            if len(dates) == 2:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = datetime.strptime(dates[1].strip(), '%m/%Y').date()
                # Ajustar end_date para que sea el último día del mes
                end_date = end_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y') + " - " + end_date.strftime('%m/%Y')
            elif len(dates) == 1:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                # Ajustar end_date para que sea el último día del mes
                end_date = start_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y')
            else:
                start_date = end_date = None
            return start_date, end_date, mes_anio_excel
        except (ValueError, IndexError) as e:
            print(f"Error processing date range: {daterange} - {e}")
            return None, None, None

    # Procesar el rango de fechas
    if daterangemayor:
        start_date, end_date, mes_anio_excel = parse_daterange(daterangemayor)

    # Si no se proporcionó un rango de fechas, usar el mes actual
    if not mes_anio_excel:
        fecha_actual = datetime.now()
        mes_anio_excel = fecha_actual.strftime('%m/%Y')
        start_date = fecha_actual.replace(day=1).date()
        end_date = (fecha_actual.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
    
    # Obtener datos del libro mayor con glosa agrupados por fecha y filtro de fecha
    libro_mayor = obtener_libro_mayor_agrupado_por_fecha_y_glosa_unica(start_date, end_date)

    # Verificar si hay datos
    if not libro_mayor:
        response = make_response(jsonify({'error': 'No hay datos para exportar en el rango de fechas seleccionado.'}), 200)
        response.headers['Content-Type'] = 'application/json'
        return response

    # Cargar el archivo de plantilla
    workbook = openpyxl.load_workbook(template_path)
    template_sheet = workbook.active

    for codigo_cuenta, detalles in libro_mayor.items():
        # Crear una copia de la hoja de plantilla
        worksheet = workbook.copy_worksheet(template_sheet)
        worksheet.title = str(codigo_cuenta)

        # Llenar las celdas específicas con los datos necesarios
        worksheet['B3'] = mes_anio_excel  # Coloca el período en la celda derecha de 'PERÍODO:'
        worksheet['B4'] = '20610588981'    # Coloca el RUC
        worksheet['B5'] = 'Tormenta'       # Coloca la razón social
        worksheet['B6'] = codigo_cuenta    # Coloca el código de la cuenta
        
        # Ajustar la anchura de la columna "C"
        worksheet.column_dimensions['C'].width = 50  # Puedes ajustar este valor según tus necesidades

        # Definir el estilo de fuente y alineación
        font = openpyxl.styles.Font(size=10, bold=False)
        alignment_left = Alignment(horizontal='left')
        alignment_center = Alignment(horizontal='center')  # Alineación centrada

        # Insertar los detalles en la tabla, fila por fila
        start_row = 11
        current_row = start_row
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for detalle in detalles:
            worksheet[f'A{current_row}'] = detalle['fecha_asiento'].strftime('%d/%m/%Y')
            worksheet[f'A{current_row}'].font = font
            worksheet[f'A{current_row}'].border = thin_border


            worksheet[f'C{current_row}'] = detalle['glosa']
            worksheet[f'C{current_row}'].font = font
            worksheet[f'C{current_row}'].alignment = alignment_left  # Alinear a la izquierda
            worksheet[f'C{current_row}'].border = thin_border
            worksheet[f'A{current_row}'].alignment = alignment_center  # Alinear al centro

            saldo_debe = detalle['total_debe']
            saldo_haber = detalle['total_haber']

            worksheet[f'D{current_row}'] = saldo_debe if saldo_debe > 0 else None
            worksheet[f'D{current_row}'].number_format = '0.00#################'
            worksheet[f'D{current_row}'].font = font
            worksheet[f'D{current_row}'].border = thin_border

            worksheet[f'E{current_row}'] = saldo_haber if saldo_haber > 0 else None
            worksheet[f'E{current_row}'].number_format = '0.00#################'
            worksheet[f'E{current_row}'].font = font
            worksheet[f'E{current_row}'].border = thin_border

            current_row += 1

        # Añadir el total al final de la tabla
        worksheet[f'C{current_row}'] = "Total"
        worksheet[f'C{current_row}'].font = openpyxl.styles.Font(size=10, bold=True)
        worksheet[f'C{current_row}'].alignment = alignment_left  # Alinear a la izquierda
        worksheet[f'C{current_row}'].border = thin_border

        # Sumar las columnas D (Deudor) y E (Acreedor)
        total_debe_formula = f'=SUM(D{start_row}:D{current_row - 1})'
        total_haber_formula = f'=SUM(E{start_row}:E{current_row - 1})'

        worksheet[f'D{current_row}'] = total_debe_formula
        worksheet[f'D{current_row}'].number_format = '0.00'
        worksheet[f'D{current_row}'].font = openpyxl.styles.Font(size=10, bold=True)
        worksheet[f'D{current_row}'].border = thin_border

        worksheet[f'E{current_row}'] = total_haber_formula
        worksheet[f'E{current_row}'].number_format = '0.00'
        worksheet[f'E{current_row}'].font = openpyxl.styles.Font(size=10, bold=True)
        worksheet[f'E{current_row}'].border = thin_border

    # Eliminar la hoja de plantilla original si no la necesitas
    workbook.remove(template_sheet)

    # Guardar el archivo modificado en un buffer de memoria
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Enviar el archivo generado
    return send_file(output, download_name="libro_mayor.xlsx", as_attachment=True)


@accounting_bp.route('/exportar_libro_diario_pdf', methods=['GET'])
@jwt_required()
def exportar_libro_diario_pdf():
    from fpdf import FPDF
    from flask import send_file
    from datetime import datetime
    from io import BytesIO

    # Crear un objeto FPDF en orientación horizontal
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    # Encabezado
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(0, 10, 'Libro Diario', ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, f'Período: {datetime.now().strftime("%m/%Y")}', ln=True, align='C')
    pdf.cell(0, 10, 'RUC: 20610588981', ln=True, align='C')
    pdf.cell(0, 10, 'Razón Social: Tormenta', ln=True, align='C')

    pdf.ln(10)  # Espacio vertical

    # Encabezados de la tabla
    pdf.set_font("Arial", style='B', size=10)
    pdf.cell(20, 10, 'N°', border=1, align='C')
    pdf.cell(30, 10, 'Fecha', border=1, align='C')
    pdf.cell(100, 10, 'Glosa', border=1, align='C')  # Aumentar el ancho de la columna Glosa
    pdf.cell(30, 10, 'Documento', border=1, align='C')
    pdf.cell(35, 10, 'Debe', border=1, align='C')
    pdf.cell(35, 10, 'Haber', border=1, align='C')
    pdf.ln(10)

    # Datos de los asientos
    asientos, _ = obtener_asientos_agrupados()
    numero_correlativo = 1
    total_debe = 0
    total_haber = 0

    pdf.set_font("Arial", size=10)
    for id_asiento, asiento in asientos.items():
        for detalle in asiento['detalles']:
            pdf.cell(20, 10, str(numero_correlativo), border=1, align='C')
            pdf.cell(30, 10, asiento['fecha_asiento'].strftime('%d/%m/%Y'), border=1, align='C')
            
            # Reducir el tamaño de la fuente para la glosa
            pdf.set_font("Arial", size=8)
            pdf.cell(100, 10, asiento['glosa'], border=1)
            pdf.set_font("Arial", size=10)  # Restaurar el tamaño de la fuente
            
            pdf.cell(30, 10, asiento['num_comprobante'], border=1, align='C')
            pdf.cell(35, 10, f"{detalle['debe']:.2f}" if detalle['debe'] != 0 else '', border=1, align='R')
            pdf.cell(35, 10, f"{detalle['haber']:.2f}" if detalle['haber'] != 0 else '', border=1, align='R')
            pdf.ln(10)
            
            total_debe += detalle['debe']
            total_haber += detalle['haber']
            numero_correlativo += 1

    # Totales
    pdf.set_font("Arial", style='B', size=10)
    pdf.cell(180, 10, 'Total', border=1, align='R')
    pdf.cell(35, 10, f"{total_debe:.2f}", border=1, align='R')
    pdf.cell(35, 10, f"{total_haber:.2f}", border=1, align='R')

    # Guardar el archivo PDF en un buffer de memoria
    output = BytesIO()
    pdf.output(dest='S').encode('latin1')  # Generar el contenido del PDF como bytes
    output.write(pdf.output(dest='S').encode('latin1'))
    output.seek(0)

    return send_file(output, download_name="libro_diario.pdf", as_attachment=True)


@accounting_bp.route('/exportar_libro_mayor_pdf', methods=['GET'])
@jwt_required()
def exportar_libro_mayor_pdf():
    # Obtener datos del libro mayor con glosa agrupados por fecha
    libro_mayor = obtener_libro_mayor_agrupado_por_fecha_y_glosa_unica()

    # Obtener la fecha actual para el período
    fecha_actual = datetime.now()
    mes_anio_actual = fecha_actual.strftime('%m/%Y')

    pdf_files = []

    for codigo_cuenta, detalles in libro_mayor.items():
        pdf = FPDF(orientation='L')
        pdf.add_page()

        # Configurar la fuente
        pdf.set_font("Arial", size=10)

        # Título
        pdf.set_font("Arial", style='B', size=14)
        pdf.cell(0, 10, "FORMATO 6.1: \"LIBRO MAYOR\"", ln=True, align='C')

        # Información del período, RUC, razón social y cuenta
        pdf.set_font("Arial", size=10)
        pdf.cell(40, 10, f"PERÍODO: {mes_anio_actual}", ln=True)
        pdf.cell(40, 10, "RUC: 20610588981", ln=True)
        pdf.cell(40, 10, "APELLIDOS Y NOMBRE: Tormenta", ln=True)
        pdf.cell(40, 10, f"CODIGO Y/O DENOMINACIÓN: {codigo_cuenta}", ln=True)

        # Encabezado de tabla
        pdf.set_font("Arial", style='B', size=10)
        pdf.cell(60, 10, "FECHA DE LA OPERACIÓN", border=1)
        pdf.cell(100, 10, "DESCRIPCIÓN O GLOSA", border=1)
        pdf.cell(50, 10, "DEUDOR", border=1, align='R')
        pdf.cell(50, 10, "ACREEDOR", border=1, align='R')
        pdf.ln()

        # Detalles del libro mayor
        total_debe = 0
        total_haber = 0

        pdf.set_font("Arial", size=10)
        for detalle in detalles:
            fecha = detalle['fecha_asiento'].strftime('%d/%m/%Y')
            glosa = detalle['glosa']
            debe = detalle['total_debe']
            haber = detalle['total_haber']

            pdf.cell(60, 10, fecha, border=1)
            pdf.cell(100, 10, glosa, border=1)
            pdf.cell(50, 10, f"{debe:.2f}", border=1, align='R')
            pdf.cell(50, 10, f"{haber:.2f}", border=1, align='R')
            pdf.ln()

            total_debe += debe
            total_haber += haber

        # Total al final de la tabla
        pdf.set_font("Arial", style='B', size=10)
        pdf.cell(160, 10, "Total", border=1, align='R')
        pdf.cell(50, 10, f"{total_debe:.2f}", border=1, align='R')
        pdf.cell(50, 10, f"{total_haber:.2f}", border=1, align='R')
        pdf.ln()

        # Guardar el PDF en memoria
        output = BytesIO()
        output.write(pdf.output(dest='S').encode('latin1'))
        output.seek(0)

        pdf_files.append({
            'filename': f"libro_mayor_{codigo_cuenta}.pdf",
            'content': output.getvalue()
        })

    # Enviar todos los archivos generados
    if len(pdf_files) == 1:
        return send_file(BytesIO(pdf_files[0]['content']), download_name=pdf_files[0]['filename'], as_attachment=True)
    else:
        import zipfile
        zip_output = BytesIO()
        with zipfile.ZipFile(zip_output, 'w') as zf:
            for archivo in pdf_files:
                zf.writestr(archivo['filename'], archivo['content'])
        zip_output.seek(0)
        return send_file(zip_output, download_name="libro_mayor.zip", as_attachment=True)
    
@accounting_bp.route('/exportar_registro_ventas_excel', methods=['GET'])
@jwt_required()
def exportar_registro_ventas_excel():
    # Ruta de la plantilla
    template_path = os.path.join(current_app.root_path, 'templates', 'contable', 'plantillas', 'Registro_Ventas.xlsx')
    
    if not os.path.exists(template_path):
        return jsonify({'error': 'La plantilla de Excel no se encontró.'}), 404
    
    # Capturar filtros desde la URL
    daterange = request.args.get('daterange', '')
    
    start_date, end_date = None, None
    mes_anio_excel = None  # Variable para guardar el valor del período en B3

    thin_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # Procesar el rango de fechas
    def parse_daterange(daterange):
        try:
            daterange_clean = re.sub(r'\s+', ' ', daterange.strip())
            dates = re.split(r'\s*(to|a)\s*', daterange_clean)
            dates = [d for d in dates if d not in ('to', 'a')]
            if len(dates) == 2:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = datetime.strptime(dates[1].strip(), '%m/%Y').date()
                end_date = end_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y') + " - " + end_date.strftime('%m/%Y')
            elif len(dates) == 1:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = start_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y')
            else:
                start_date = end_date = None
            return start_date, end_date, mes_anio_excel
        except (ValueError, IndexError):
            return None, None, None

    if daterange:
        start_date, end_date, mes_anio_excel = parse_daterange(daterange)

    # Si no se proporcionó rango de fechas, usar el mes actual
    if not mes_anio_excel:
        fecha_actual = datetime.now()
        mes_anio_excel = fecha_actual.strftime('%m/%Y')
        start_date = fecha_actual.replace(day=1).date()
        end_date = (fecha_actual.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
    
    print(start_date)
    print(end_date)
    # Obtener datos del registro de ventas
    registros_ventas, totales = obtener_registro_ventas(start_date, end_date)

    # Verificar si hay datos
    if not registros_ventas:
        response = make_response(jsonify({'error': 'No hay datos para exportar en el rango de fechas seleccionado.'}), 200)
        response.headers['Content-Type'] = 'application/json'
        return response

    # Cargar el archivo de plantilla
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Rellenar encabezados
    worksheet['B3'] = mes_anio_excel
    worksheet['B4'] = '20610588981'  # RUC
    worksheet['B5'] = 'Tormenta'     # Razón social

    # Iniciar desde la fila 11
    start_row = 12
    current_row = start_row

    # Estilo de fuente
    font_style = Font(name='Calibri', size=11)

    # Alineación
    alignment_center = Alignment(horizontal='center')
    alignment_left = Alignment(horizontal='left')
    alignment_right = Alignment(horizontal='right')

    for registro in registros_ventas:
        worksheet[f'A{current_row}'] = registro["numero_correlativo"]
        worksheet[f'A{current_row}'].border = thin_border
        worksheet[f'A{current_row}'].font = font_style
        worksheet[f'A{current_row}'].alignment = alignment_center

        worksheet[f'B{current_row}'] = registro["fecha"].strftime('%d/%m/%Y')
        worksheet[f'B{current_row}'].border = thin_border
        worksheet[f'B{current_row}'].font = font_style
        worksheet[f'B{current_row}'].alignment = alignment_center

        worksheet[f'C{current_row}'] = registro["fechaV"].strftime('%d/%m/%Y')
        worksheet[f'C{current_row}'].border = thin_border
        worksheet[f'C{current_row}'].font = font_style
        worksheet[f'C{current_row}'].alignment = alignment_center

        worksheet[f'D{current_row}'] = '01' if str(registro["num_comprobante"])[0] == 'F' else '03'
        worksheet[f'D{current_row}'].border = thin_border
        worksheet[f'D{current_row}'].font = font_style
        worksheet[f'D{current_row}'].alignment = alignment_center

        worksheet[f'E{current_row}'] = registro["num_comprobante"].split('-')[0][1:]
        worksheet[f'E{current_row}'].border = thin_border
        worksheet[f'E{current_row}'].font = font_style
        worksheet[f'E{current_row}'].alignment = alignment_center

        worksheet[f'F{current_row}'] = registro["num_comprobante"].split('-')[1]
        worksheet[f'F{current_row}'].border = thin_border
        worksheet[f'F{current_row}'].font = font_style
        worksheet[f'F{current_row}'].alignment = alignment_center

        worksheet[f'G{current_row}'] = '1' if len(str(registro["documento_cliente"])) == 8 else '6'
        worksheet[f'G{current_row}'].border = thin_border
        worksheet[f'G{current_row}'].font = font_style
        worksheet[f'G{current_row}'].alignment = alignment_center

        worksheet[f'H{current_row}'] = registro["documento_cliente"]
        worksheet[f'H{current_row}'].border = thin_border
        worksheet[f'H{current_row}'].font = font_style
        worksheet[f'H{current_row}'].alignment = alignment_center

        worksheet[f'I{current_row}'] = registro["nombre_cliente"]
        worksheet[f'I{current_row}'].border = thin_border
        worksheet[f'I{current_row}'].font = font_style
        worksheet[f'I{current_row}'].alignment = alignment_left

        worksheet[f'K{current_row}'] = registro["importe"]
        worksheet[f'K{current_row}'].border = thin_border
        worksheet[f'K{current_row}'].font = font_style
        worksheet[f'K{current_row}'].alignment = alignment_right
        worksheet[f'K{current_row}'].number_format = '#,##0.00'

        worksheet[f'O{current_row}'] = registro["igv"]
        worksheet[f'O{current_row}'].border = thin_border
        worksheet[f'O{current_row}'].font = font_style
        worksheet[f'O{current_row}'].alignment = alignment_right
        worksheet[f'O{current_row}'].number_format = '#,##0.00'

        worksheet[f'Q{current_row}'] = registro["total"]
        worksheet[f'Q{current_row}'].border = thin_border
        worksheet[f'Q{current_row}'].font = font_style
        worksheet[f'Q{current_row}'].alignment = alignment_right
        worksheet[f'Q{current_row}'].number_format = '#,##0.00'

        current_row += 1


    # Añadir totales al final
    total_row = current_row
    worksheet.merge_cells(f'H{total_row}:J{total_row}')
    worksheet[f'H{total_row}'] = 'Totales'
    worksheet[f'H{total_row}'].border = thin_border
    worksheet[f'H{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'H{total_row}'].alignment = alignment_right

    worksheet[f'K{total_row}'] = totales["total_importe"]
    worksheet[f'K{total_row}'].border = thin_border
    worksheet[f'K{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'K{total_row}'].alignment = alignment_right
    worksheet[f'K{total_row}'].number_format = '#,##0.00'

    worksheet[f'O{total_row}'] = totales["total_igv"]
    worksheet[f'O{total_row}'].border = thin_border
    worksheet[f'O{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'O{total_row}'].alignment = alignment_right
    worksheet[f'O{total_row}'].number_format = '#,##0.00'

    worksheet[f'Q{total_row}'] = totales["total_general"]
    worksheet[f'Q{total_row}'].border = thin_border
    worksheet[f'Q{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'Q{total_row}'].alignment = alignment_right
    worksheet[f'Q{total_row}'].number_format = '#,##0.00'

    # Guardar el archivo modificado en un buffer
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Enviar el archivo generado
    return send_file(output, download_name="registro_ventas.xlsx", as_attachment=True)

@accounting_bp.route('/exportar_registro_ventas_pdf', methods=['GET'])
@jwt_required()
def exportar_registro_ventas_pdf():
    from fpdf import FPDF
    from flask import send_file
    from datetime import datetime
    from io import BytesIO

    # Crear un objeto FPDF en orientación horizontal
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    # Encabezado
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(0, 10, 'Registro de Ventas', ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, f'Período: {datetime.now().strftime("%m/%Y")}', ln=True, align='C')
    pdf.cell(0, 10, 'RUC: 20610588981', ln=True, align='C')
    pdf.cell(0, 10, 'Razón Social: Tormenta', ln=True, align='C')

    pdf.ln(10)  # Espacio vertical

    # Encabezados de la tabla
    pdf.set_font("Arial", style='B', size=9)
    headers = [
        ('N°', 10), ('F. Emisión', 23), ('F. Vencimiento', 25), 
        ('Tipo', 10), ('Serie', 13), ('Número', 20), ('Tipo Doc.', 15), 
        ('N° Doc', 25), ('Nombre Cliente', 70), 
        ('Importe', 20), ('IGV', 20), ('Total', 20)
    ]
    for header, width in headers:
        pdf.cell(width, 10, header, border=1, align='C')
    pdf.ln(10)

    # Obtener registros de ventas
    registros_ventas, totales = obtener_registro_ventas()
    
    pdf.set_font("Arial", size=9)
    for registro in registros_ventas:
        pdf.cell(10, 10, str(registro["numero_correlativo"]), border=1, align='C')
        pdf.cell(23, 10, registro["fecha"].strftime('%d/%m/%Y'), border=1, align='C')
        pdf.cell(25, 10, registro["fechaV"].strftime('%d/%m/%Y'), border=1, align='C')
        
        tipo_documento = '01' if registro["num_comprobante"][0] == "F" else '03' if registro["num_comprobante"][0] == "B" else '07'
        pdf.cell(10, 10, tipo_documento, border=1, align='C')

        
        num_comprobante = registro["num_comprobante"]
        serie = num_comprobante.split("-")[0][1:]  #Serie
        numero = num_comprobante.split("-")[1]      # Numero
        
        pdf.cell(13, 10, serie, border=1, align='C')   # Celda para la serie
        pdf.cell(20, 10, numero, border=1, align='C')  # Celda para el número
        
        # Tipo de documento del cliente
        tipo_doc_cliente = '1' if len(str(registro["documento_cliente"])) == 8 else '6'
        pdf.cell(15, 10, tipo_doc_cliente, border=1, align='C')
        
        pdf.cell(25, 10, str(registro["documento_cliente"]), border=1, align='C')
        pdf.cell(70, 10, registro["nombre_cliente"], border=1)
        pdf.cell(20, 10, f"{registro['importe']:.2f}", border=1, align='R')
        pdf.cell(20, 10, f"{registro['igv']:.2f}", border=1, align='R')
        pdf.cell(20, 10, f"{registro['total']:.2f}", border=1, align='R')
        pdf.ln(10)

    # Totales
    pdf.set_font("Arial", style='B', size=10)
    pdf.cell(211, 10, 'Totales', border=1, align='R')
    pdf.cell(20, 10, f"{totales['total_importe']:.2f}", border=1, align='R')
    pdf.cell(20, 10, f"{totales['total_igv']:.2f}", border=1, align='R')
    pdf.cell(20, 10, f"{totales['total_general']:.2f}", border=1, align='R')

    # Guardar el archivo PDF en un buffer de memoria
    output = BytesIO()
    pdf.output(dest='S').encode('latin1')
    output.write(pdf.output(dest='S').encode('latin1'))
    output.seek(0)

    return send_file(output, download_name="registro_ventas.pdf", as_attachment=True)


@accounting_bp.route('/exportar_registro_compras_excel', methods=['GET'])
@jwt_required()
def exportar_registro_compras_excel():
    # Ruta de la plantilla
    template_path = os.path.join(current_app.root_path, 'templates', 'contable', 'plantillas', 'Registro_Compras.xlsx')
    
    if not os.path.exists(template_path):
        return jsonify({'error': 'La plantilla de Excel no se encontró.'}), 404
    
    # Capturar filtros desde la URL
    daterange = request.args.get('daterange', '')
    
    start_date, end_date = None, None
    mes_anio_excel = None  # Variable para guardar el valor del período en B3

    # Procesar el rango de fechas
    def parse_daterange(daterange):
        try:
            daterange_clean = re.sub(r'\s+', ' ', daterange.strip())
            dates = re.split(r'\s*(to|a)\s*', daterange_clean)
            dates = [d for d in dates if d not in ('to', 'a')]
            if len(dates) == 2:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = datetime.strptime(dates[1].strip(), '%m/%Y').date()
                end_date = end_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y') + " - " + end_date.strftime('%m/%Y')
            elif len(dates) == 1:
                start_date = datetime.strptime(dates[0].strip(), '%m/%Y').date()
                end_date = start_date.replace(day=1) + timedelta(days=31)
                end_date = end_date.replace(day=1) - timedelta(days=1)
                mes_anio_excel = start_date.strftime('%m/%Y')
            else:
                start_date = end_date = None
            return start_date, end_date, mes_anio_excel
        except (ValueError, IndexError):
            return None, None, None

    if daterange:
        start_date, end_date, mes_anio_excel = parse_daterange(daterange)

    # Si no se proporcionó rango de fechas, usar el mes actual
    if not mes_anio_excel:
        fecha_actual = datetime.now()
        mes_anio_excel = fecha_actual.strftime('%m/%Y')
        start_date = fecha_actual.replace(day=1).date()
        end_date = (fecha_actual.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
    
    # Obtener datos del registro de compras
    registros_compras, totales = obtener_registro_compras(start_date, end_date)

    # Verificar si hay datos
    if not registros_compras:
        response = make_response(jsonify({'error': 'No hay datos para exportar en el rango de fechas seleccionado.'}), 200)
        response.headers['Content-Type'] = 'application/json'
        return response

    # Cargar el archivo de plantilla
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Rellenar encabezados
    worksheet['B3'] = mes_anio_excel
    worksheet['B4'] = '20610588981'  # RUC
    worksheet['B5'] = 'Tormenta'     # Razón social

    # Definir estilos
    thin_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    font_style = Font(name='Calibri', size=11)
    alignment_center = Alignment(horizontal='center')
    alignment_left = Alignment(horizontal='left')
    alignment_right = Alignment(horizontal='right')

    # Aplicar bordes y estilos a los encabezados
    header_row = 13
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        cell = worksheet[f'{col}{header_row}']
        cell.border = thin_border
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = alignment_center

    # Ajustar ancho de columnas
    column_widths = {
        'A': 10,
        'B': 12,
        'C': 12,
        'D': 5,
        'E': 8,
        'F': 8,
        'G': 10,
        'H': 5,
        'I': 15,
        'J': 30,
        'K': 15,
        'L': 15,
        'M': 15,
        # Agrega más columnas si es necesario
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # Iniciar desde la fila 14
    start_row = 14
    current_row = start_row

    for registro in registros_compras:
        worksheet[f'A{current_row}'] = registro["numero_correlativo"]  # Número correlativo
        worksheet[f'A{current_row}'].border = thin_border
        worksheet[f'A{current_row}'].font = font_style
        worksheet[f'A{current_row}'].alignment = alignment_center

        worksheet[f'B{current_row}'] = registro["fecha"].strftime('%d/%m/%Y')  # Fecha de emisión
        worksheet[f'B{current_row}'].border = thin_border
        worksheet[f'B{current_row}'].font = font_style
        worksheet[f'B{current_row}'].alignment = alignment_center

        worksheet[f'C{current_row}'] = registro["fechaV"].strftime('%d/%m/%Y')  # Fecha de vencimiento
        worksheet[f'C{current_row}'].border = thin_border
        worksheet[f'C{current_row}'].font = font_style
        worksheet[f'C{current_row}'].alignment = alignment_center

        # Lógica para tipo de comprobante basado en el número
        if str(registro["num_comprobante"])[0] == "F":
            worksheet[f'D{current_row}'] = '01'  # Tipo
        elif str(registro["num_comprobante"])[0] == "B":
            worksheet[f'D{current_row}'] = '03'  # Tipo
        else:
            worksheet[f'D{current_row}'] = '07'  # Tipo
        worksheet[f'D{current_row}'].border = thin_border
        worksheet[f'D{current_row}'].font = font_style
        worksheet[f'D{current_row}'].alignment = alignment_center

        # Dividir el número de comprobante en serie y número
        comprobante = registro["num_comprobante"]
        serie = comprobante.split("-")[0][1:]
        numero = comprobante.split("-")[1]

        worksheet[f'E{current_row}'] = serie  # Serie
        worksheet[f'E{current_row}'].border = thin_border
        worksheet[f'E{current_row}'].font = font_style
        worksheet[f'E{current_row}'].alignment = alignment_center

        worksheet[f'F{current_row}'] = ''  # Se deja vacío según tu código original
        worksheet[f'F{current_row}'].border = thin_border
        worksheet[f'F{current_row}'].font = font_style
        worksheet[f'F{current_row}'].alignment = alignment_center

        worksheet[f'G{current_row}'] = numero  # Número
        worksheet[f'G{current_row}'].border = thin_border
        worksheet[f'G{current_row}'].font = font_style
        worksheet[f'G{current_row}'].alignment = alignment_center

        # Tipo de documento del proveedor basado en el número de documento
        if len(str(registro["documento_cliente"])) == 8:
            worksheet[f'H{current_row}'] = '1' 
        else:
            worksheet[f'H{current_row}'] = '6'  # Tipo
        worksheet[f'H{current_row}'].border = thin_border
        worksheet[f'H{current_row}'].font = font_style
        worksheet[f'H{current_row}'].alignment = alignment_center

        worksheet[f'I{current_row}'] = registro["documento_cliente"]  # Número de documento
        worksheet[f'I{current_row}'].border = thin_border
        worksheet[f'I{current_row}'].font = font_style
        worksheet[f'I{current_row}'].alignment = alignment_center

        worksheet[f'J{current_row}'] = registro["nombre_cliente"]  # Razón social
        worksheet[f'J{current_row}'].border = thin_border
        worksheet[f'J{current_row}'].font = font_style
        worksheet[f'J{current_row}'].alignment = alignment_left

        worksheet[f'K{current_row}'] = registro["importe"]  # Importe
        worksheet[f'K{current_row}'].border = thin_border
        worksheet[f'K{current_row}'].font = font_style
        worksheet[f'K{current_row}'].alignment = alignment_right
        worksheet[f'K{current_row}'].number_format = '#,##0.00'

        worksheet[f'L{current_row}'] = registro["igv"]  # IGV
        worksheet[f'L{current_row}'].border = thin_border
        worksheet[f'L{current_row}'].font = font_style
        worksheet[f'L{current_row}'].alignment = alignment_right
        worksheet[f'L{current_row}'].number_format = '#,##0.00'

        worksheet[f'M{current_row}'] = registro["total"]  # Total
        worksheet[f'M{current_row}'].border = thin_border
        worksheet[f'M{current_row}'].font = font_style
        worksheet[f'M{current_row}'].alignment = alignment_right
        worksheet[f'M{current_row}'].number_format = '#,##0.00'

        # Si tienes otras columnas (N, O, P, etc.), aplica el mismo procedimiento

        current_row += 1

    # Añadir totales al final
    total_row = current_row
    worksheet.merge_cells(f'H{total_row}:I{total_row}')
    worksheet[f'H{total_row}'] = 'Totales'
    worksheet[f'H{total_row}'].border = thin_border
    worksheet[f'H{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'H{total_row}'].alignment = alignment_right

    worksheet[f'K{total_row}'] = totales["total_importe"]
    worksheet[f'K{total_row}'].border = thin_border
    worksheet[f'K{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'K{total_row}'].alignment = alignment_right
    worksheet[f'K{total_row}'].number_format = '#,##0.00'

    worksheet[f'L{total_row}'] = totales["total_igv"]
    worksheet[f'L{total_row}'].border = thin_border
    worksheet[f'L{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'L{total_row}'].alignment = alignment_right
    worksheet[f'L{total_row}'].number_format = '#,##0.00'

    worksheet[f'M{total_row}'] = totales["total_general"]
    worksheet[f'M{total_row}'].border = thin_border
    worksheet[f'M{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    worksheet[f'M{total_row}'].alignment = alignment_right
    worksheet[f'M{total_row}'].number_format = '#,##0.00'

    # Guardar el archivo modificado en un buffer
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Enviar el archivo generado
    return send_file(output, download_name="registro_compras.xlsx", as_attachment=True)

@accounting_bp.route('/exportar_registro_compras_pdf', methods=['GET'])
@jwt_required()
def exportar_registro_compras_pdf():
    from fpdf import FPDF
    from flask import send_file
    from datetime import datetime
    from io import BytesIO

    # Crear un objeto FPDF en orientación horizontal
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    # Encabezado
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(0, 10, 'Registro de Compras', ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, f'Período: {datetime.now().strftime("%m/%Y")}', ln=True, align='C')
    pdf.cell(0, 10, 'RUC: 20610588981', ln=True, align='C')
    pdf.cell(0, 10, 'Razón Social: Tormenta', ln=True, align='C')

    pdf.ln(10)  # Espacio vertical

    # Encabezados de la tabla
    pdf.set_font("Arial", style='B', size=9)
    headers = [
        ('N°', 10), ('F. Emisión', 23), ('F. Vencimiento', 25), 
        ('Tipo', 10), ('Serie', 13), ('Número', 20), ('Tipo Doc.', 15), 
        ('N° Doc', 25), ('Razón Social', 70), 
        ('Importe', 20), ('IGV', 20), ('Total', 20)
    ]
    for header, width in headers:
        pdf.cell(width, 10, header, border=1, align='C')
    pdf.ln(10)

    # Obtener registros de compras
    registros_compras, totales = obtener_registro_compras()
    
    pdf.set_font("Arial", size=9)
    for registro in registros_compras:
        pdf.cell(10, 10, str(registro["numero_correlativo"]), border=1, align='C')
        pdf.cell(23, 10, registro["fecha"].strftime('%d/%m/%Y'), border=1, align='C')
        pdf.cell(25, 10, registro["fechaV"].strftime('%d/%m/%Y'), border=1, align='C')
        
        # Determinar el tipo de comprobante
        tipo_documento = '01' if registro["num_comprobante"][0] == "F" else '03' if registro["num_comprobante"][0] == "B" else '07'
        pdf.cell(10, 10, tipo_documento, border=1, align='C')

        # Serie y número del comprobante
        num_comprobante = registro["num_comprobante"]
        serie = num_comprobante.split("-")[0][1:]  # Serie
        numero = num_comprobante.split("-")[1]      # Número
        
        pdf.cell(13, 10, serie, border=1, align='C')   # Celda para la serie
        pdf.cell(20, 10, numero, border=1, align='C')  # Celda para el número

        # Tipo de documento del cliente
        tipo_doc_cliente = '1' if len(str(registro["documento_cliente"])) == 8 else '6'
        pdf.cell(15, 10, tipo_doc_cliente, border=1, align='C')
        
        pdf.cell(25, 10, str(registro["documento_cliente"]), border=1, align='C')
        pdf.cell(70, 10, registro["nombre_cliente"], border=1)
        pdf.cell(20, 10, f"{registro['importe']:.2f}", border=1, align='R')
        pdf.cell(20, 10, f"{registro['igv']:.2f}", border=1, align='R')
        pdf.cell(20, 10, f"{registro['total']:.2f}", border=1, align='R')
        pdf.ln(10)

    # Totales
    pdf.set_font("Arial", style='B', size=10)
    pdf.cell(211, 10, 'Totales', border=1, align='R')
    pdf.cell(20, 10, f"{totales['total_importe']:.2f}", border=1, align='R')
    pdf.cell(20, 10, f"{totales['total_igv']:.2f}", border=1, align='R')
    pdf.cell(20, 10, f"{totales['total_general']:.2f}", border=1, align='R')

    # Guardar el archivo PDF en un buffer de memoria
    output = BytesIO()
    pdf.output(dest='S').encode('latin1')
    output.write(pdf.output(dest='S').encode('latin1'))
    output.seek(0)

    return send_file(output, download_name="registro_compras.pdf", as_attachment=True)

from calendar import monthrange

@accounting_bp.route('/exportar_libro_caja_excel', methods=['GET'])
def exportar_libro_caja_excel():
    # Define la ruta de la plantilla de Excel
    template_path = os.path.join(current_app.root_path, 'templates', 'contable', 'plantillas', 'libro_caja.xlsx')
    
    # Verifica si la plantilla existe
    if not os.path.exists(template_path):
        return "La plantilla de Excel no se encontró.", 404
    
    output = BytesIO()
    workbook = load_workbook(template_path)
    worksheet = workbook.active

    # Obtener el rango de fechas desde los parámetros de la solicitud
    daterange_caja = request.args.get('daterangecaja', '').strip()

    if daterange_caja:
        try:
            # Parsear el mes y año seleccionados
            selected_date = datetime.strptime(daterange_caja, '%m/%Y')
            selected_month = selected_date.month
            selected_year = selected_date.year

            # Calcular las fechas de inicio y fin del mes seleccionado
            start_date_caja = datetime(selected_year, selected_month, 1).date()
            last_day = monthrange(selected_year, selected_month)[1]
            end_date_caja = datetime(selected_year, selected_month, last_day).date()
        except ValueError:
            return "Fecha inválida", 400
    else:
        # Si no se proporciona una fecha, usar el mes actual
        fecha_actual = datetime.now()
        selected_month = fecha_actual.month
        selected_year = fecha_actual.year
        start_date_caja = fecha_actual.replace(day=1).date()
        end_date_caja = last_day_of_month(fecha_actual).date()

    # Llamar a obtener_libro_caja con las fechas
    lista_libro_caja, total_caja = obtener_libro_caja(start_date_caja, end_date_caja)

    # Verificar si hay datos
    if not lista_libro_caja:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'No se encontraron registros para el criterio seleccionado.'}), 400
        else:
            return "No se encontraron registros para el criterio seleccionado.", 404

    # Formatear mes y año para el encabezado
    mes_anio_actual = f"{selected_month:02d}/{selected_year}"

    # Llenar las celdas con los datos de encabezado
    worksheet['B3'] = mes_anio_actual
    worksheet['B4'] = '20610588981'  # Ejemplo de RUC
    worksheet['B5'] = 'Tormenta'  # Nombre de la empresa

    # Define la fila inicial para los datos
    start_row = 9

    # Borde delgado para las celdas
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Estilo de fuente y alineación
    normal_font = Font(bold=False)
    right_alignment = Alignment(horizontal='right', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Variable para el número correlativo
    numero_correlativo = 1

    # Itera sobre los datos y llénalos en el Excel
    current_row = start_row
    for row in lista_libro_caja:
        worksheet[f'A{current_row}'] = numero_correlativo
        worksheet[f'B{current_row}'] = row['fecha'].strftime('%d/%m/%Y')
        worksheet[f'C{current_row}'] = row['glosa']
        worksheet[f'D{current_row}'] = row['cod_cuenta']
        worksheet[f'E{current_row}'] = row['nombre_cuenta']
        worksheet[f'F{current_row}'] = row['debe']
        worksheet[f'G{current_row}'] = row['haber']

        # Aplica bordes y alineación
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = worksheet[f'{col}{current_row}']
            cell.border = thin_border
            cell.font = normal_font

        # Alineación específica para columnas
        worksheet[f'A{current_row}'].alignment = center_alignment
        worksheet[f'B{current_row}'].alignment = center_alignment
        worksheet[f'C{current_row}'].alignment = left_alignment
        worksheet[f'D{current_row}'].alignment = center_alignment
        worksheet[f'E{current_row}'].alignment = left_alignment
        worksheet[f'F{current_row}'].alignment = right_alignment
        worksheet[f'G{current_row}'].alignment = right_alignment

        # Incrementa el número correlativo y la fila actual
        numero_correlativo += 1
        current_row += 1

    # Agrega los totales en la última fila
    total_row = current_row
    worksheet.merge_cells(f'A{total_row}:E{total_row}')
    worksheet[f'A{total_row}'] = "Total"
    worksheet[f'F{total_row}'] = total_caja['debe']
    worksheet[f'G{total_row}'] = total_caja['haber']

    # Aplica formato y bordes a los totales
    for col in ['A', 'F', 'G']:
        cell = worksheet[f'{col}{total_row}']
        cell.border = thin_border
        cell.alignment = right_alignment

    # Asegura el ancho de las columnas para mejor visibilidad
    column_widths = {'A': 38, 'B': 15, 'C': 60, 'D': 15, 'E': 30, 'F': 15, 'G': 15}
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # Guarda el archivo en el buffer y lo envía
    workbook.save(output)
    output.seek(0)
    
    return send_file(output, download_name="libro_caja.xlsx", as_attachment=True)

@accounting_bp.route('/exportar_libro_caja_pdf', methods=['GET'])
@jwt_required()
@role_required('ADMIN', 'CONTADOR')
def exportar_libro_caja_pdf():
    from fpdf import FPDF
    from io import BytesIO
    from flask import send_file, jsonify
    from calendar import monthrange

    # Obtener el rango de fechas desde los parámetros de la solicitud
    daterange_caja = request.args.get('daterangecaja', '').strip()

    if daterange_caja:
        try:
            # Parsear el mes y año seleccionados
            selected_date = datetime.strptime(daterange_caja, '%m/%Y')
            selected_month = selected_date.month
            selected_year = selected_date.year

            # Calcular las fechas de inicio y fin del mes seleccionado
            start_date_caja = datetime(selected_year, selected_month, 1).date()
            last_day = monthrange(selected_year, selected_month)[1]
            end_date_caja = datetime(selected_year, selected_month, last_day).date()
        except ValueError:
            return "Fecha inválida", 400
    else:
        # Si no se proporciona una fecha, usar el mes actual
        fecha_actual = datetime.now()
        selected_month = fecha_actual.month
        selected_year = fecha_actual.year
        start_date_caja = fecha_actual.replace(day=1).date()
        end_date_caja = last_day_of_month(fecha_actual).date()

    # Obtener datos del libro caja con el rango de fechas
    lista_libro_caja, total_caja = obtener_libro_caja(start_date_caja, end_date_caja)

    # Verificar si hay datos
    if not lista_libro_caja:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'No se encontraron registros para el criterio seleccionado.'}), 400
        else:
            return "No se encontraron registros para el criterio seleccionado.", 404

    # Crear un objeto FPDF
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Encabezado
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(0, 10, 'Libro Caja', ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, f'RUC: 20610588981', ln=True, align='C')
    pdf.cell(0, 10, 'Razón Social: Tormenta', ln=True, align='C')
    pdf.cell(0, 10, f'Periodo: {selected_month:02d}/{selected_year}', ln=True, align='C')  # Mostrar el mes y año seleccionados

    pdf.ln(10)  # Espacio vertical

    # Encabezados de la tabla
    pdf.set_font("Arial", style='B', size=10)
    headers = [
        ('N°', 10), ('Fecha', 30), ('Glosa', 105),
        ('Código', 20), ('Cuenta', 50), ('Debe', 30), ('Haber', 30)
    ]
    for header, width in headers:
        pdf.cell(width, 10, header, border=1, align='C')
    pdf.ln(10)

    # Agregar los datos
    pdf.set_font("Arial", size=9)
    numero_correlativo = 1
    for registro in lista_libro_caja:
        pdf.cell(10, 10, str(numero_correlativo), border=1, align='C')
        pdf.cell(30, 10, registro['fecha'].strftime('%d/%m/%Y'), border=1, align='C')
        pdf.cell(105, 10, registro['glosa'], border=1)
        pdf.cell(20, 10, registro['cod_cuenta'], border=1, align='C')
        pdf.cell(50, 10, registro['nombre_cuenta'], border=1)
        pdf.cell(30, 10, f"{registro['debe']:.2f}", border=1, align='R')
        pdf.cell(30, 10, f"{registro['haber']:.2f}", border=1, align='R')
        pdf.ln(10)
        numero_correlativo += 1

    # Totales
    pdf.set_font("Arial", style='B', size=10)

    # Ajustar los anchos de las columnas totales para coincidir con la tabla
    total_row_width = 10 + 30 + 105 + 20 + 50
    pdf.cell(total_row_width, 10, 'Totales', border=1, align='R')
    pdf.cell(30, 10, f"{total_caja['debe']:.2f}", border=1, align='R')
    pdf.cell(30, 10, f"{total_caja['haber']:.2f}", border=1, align='R')

    # Guardar el archivo en memoria
    output = BytesIO()
    pdf_output = pdf.output(dest='S').encode('latin1')  # Generar el contenido del PDF en bytes
    output.write(pdf_output)  # Escribir los bytes en el buffer
    output.seek(0)  # Posicionar el cursor al inicio del buffer

    # Enviar el archivo como respuesta
    return send_file(output, download_name="libro_caja.pdf", as_attachment=True)
